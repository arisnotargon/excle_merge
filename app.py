import openpyxl
import base64
import os
import re
import argparse
from copy import copy

def copy_cell(source_cell:openpyxl.cell.cell,target_cell:openpyxl.cell.cell):
    target_cell.data_type = source_cell.data_type
    target_cell.value = source_cell.value
    target_cell.fill = copy(source_cell.fill)
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)

    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)

if __name__ == "__main__":
    work_dir = 'input'
    list = os.listdir(work_dir)
    out_wb = openpyxl.load_workbook('output/06.xlsx')
    # out_sheet = out_wb[out_wb.sheetnames[0]]
    out_sheet = out_wb.active
    out_line_no = 5

    for inputfname in list:
        if inputfname[-5:] == '.xlsx':
            in_wb = openpyxl.load_workbook(work_dir + os.sep + inputfname)
            in_sheet = in_wb.active
            in_max_row_no = in_sheet.max_row
            for row_no in range(5, in_max_row_no + 1):
                judge_cell = in_sheet.cell(row_no, 1)
                if judge_cell.value != None:
                    copy_cell(in_sheet.cell(row_no, 2),out_sheet.cell(out_line_no,4))
                    copy_cell(in_sheet.cell(row_no, 3),out_sheet.cell(out_line_no,5))
                    copy_cell(in_sheet.cell(row_no, 4),out_sheet.cell(out_line_no,7))
                    copy_cell(in_sheet.cell(row_no, 5),out_sheet.cell(out_line_no,8))
                    copy_cell(in_sheet.cell(row_no, 6),out_sheet.cell(out_line_no,9))
                    copy_cell(in_sheet.cell(row_no, 7),out_sheet.cell(out_line_no,10))
                    copy_cell(in_sheet.cell(row_no, 8),out_sheet.cell(out_line_no,11))
                    copy_cell(in_sheet.cell(row_no, 9),out_sheet.cell(out_line_no,12))
                    copy_cell(in_sheet.cell(row_no, 10),out_sheet.cell(out_line_no,13))

                    out_line_no += 1

    out_wb.save('output'+os.sep+'out.xlsx')